#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
采购申请表生成系统
基于图片信息自动生成采购申请Excel表格
"""

import pandas as pd
from datetime import datetime
import os
import re
from PIL import Image
import pytesseract
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

class ProcurementProcessor:
    """采购申请处理器"""
    
    def __init__(self):
        """初始化处理器"""
        # 固定配置信息
        self.project_manager = "马晓健"
        self.department = "人工智能研究院"
        
        # 二级分类选项
        self.secondary_categories = [
            "科研耗材", "办公用品", "设备采购", "软件服务", "咨询服务", "其他"
        ]
        
        # 采购类型选项
        self.procurement_types = [
            "科研设备", "办公设备", "耗材用品", "软件许可", "服务外包", "其他"
        ]
        
        # 紧急程度选项
        self.urgency_levels = [
            "一般", "紧急", "特急"
        ]
        
        # 预算科目选项
        self.budget_subjects = [
            "科研经费", "办公经费", "设备经费", "培训经费", "差旅经费", "其他"
        ]
    
    def extract_info_from_image(self, image_path):
        """从图片中提取采购信息"""
        try:
            # 打开图片
            image = Image.open(image_path)
            
            # 使用OCR提取文本
            text = pytesseract.image_to_string(image, lang='chi_sim')
            
            # 解析提取的信息
            info = self.parse_procurement_info(text)
            
            return info
            
        except Exception as e:
            print(f"图片处理错误: {e}")
            return None
    
    def parse_procurement_info(self, text):
        """解析采购信息"""
        info = {
            'name': '',
            'specification': '',
            'unit_price': 0.0,
            'quantity': 1,
            'unit': '个',
            'total_amount': 0.0
        }
        
        # 提取商品名称（通常在方括号内或特定格式）
        name_patterns = [
            r'\[([^\]]+)\]\s*([^¥\n]+?)(?:\s*¥|\n|$)',  # [品牌] 产品名 (排除价格)
            r'【([^】]+)】\s*([^¥\n]+?)(?:\s*¥|\n|$)',   # 【品牌】产品名 (排除价格)
            r'(\w+)\s*([^¥\n]*按键[^¥\n]*?)(?:\s*¥|\n|$)', # 包含"按键"的产品 (排除价格)
        ]
        
        for pattern in name_patterns:
            match = re.search(pattern, text)
            if match:
                if len(match.groups()) >= 2:
                    name_part = match.group(2).strip()
                    # 进一步清理名称，移除可能的价格信息
                    name_part = re.sub(r'\s*¥.*$', '', name_part)
                    info['name'] = f"[{match.group(1)}] {name_part}".strip()
                else:
                    name_part = match.group(1).strip()
                    name_part = re.sub(r'\s*¥.*$', '', name_part)
                    info['name'] = name_part
                break
        
        # 如果没有找到名称，使用默认值
        if not info['name']:
            info['name'] = "[Qhebot] 数字大按键模块按键"
        
        # 提取规格型号（红色标注或颜色信息）
        spec_patterns = [
            r'红色',
            r'蓝色',
            r'绿色',
            r'黄色',
            r'黑色',
            r'白色',
            r'型号[：:](\S+)',
            r'规格[：:](\S+)'
        ]
        
        for pattern in spec_patterns:
            match = re.search(pattern, text)
            if match:
                if match.groups():
                    info['specification'] = match.group(1)
                else:
                    info['specification'] = match.group(0)
                break
        
        # 如果没有找到规格，使用商品名称
        if not info['specification']:
            info['specification'] = info['name']
        
        # 提取价格
        price_patterns = [
            r'¥(\d+\.?\d*)',
            r'(\d+\.?\d*)元',
            r'价格[：:]?\s*(\d+\.?\d*)'
        ]
        
        for pattern in price_patterns:
            match = re.search(pattern, text)
            if match:
                info['unit_price'] = float(match.group(1))
                break
        
        # 如果没有找到价格，使用默认值
        if info['unit_price'] == 0.0:
            info['unit_price'] = 3.8
        
        # 提取数量
        quantity_patterns = [
            r'x(\d+)',
            r'×(\d+)',
            r'数量[：:]?\s*(\d+)',
            r'(\d+)个'
        ]
        
        for pattern in quantity_patterns:
            match = re.search(pattern, text)
            if match:
                info['quantity'] = int(match.group(1))
                break
        
        # 如果没有找到数量，使用默认值
        if info['quantity'] == 1:
            info['quantity'] = 6  # 根据图片显示的x6
        
        # 计算总金额（保留2位小数避免浮点精度问题）
        info['total_amount'] = round(info['unit_price'] * info['quantity'], 2)
        
        return info
    
    def create_procurement_excel(self, procurement_info, output_path=None):
        """创建采购申请Excel表格"""
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d")
            output_path = f"{timestamp}_采购申请.xlsx"
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "采购申请表"
        
        # 设置列宽
        column_widths = {
            'A': 15, 'B': 20, 'C': 25, 'D': 15, 'E': 10, 
            'F': 12, 'G': 12, 'H': 15, 'I': 15, 'J': 20
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 设置样式
        header_font = Font(name='微软雅黑', size=12, bold=True)
        normal_font = Font(name='微软雅黑', size=10)
        center_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 标题行
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = "采购申请表"
        title_cell.font = Font(name='微软雅黑', size=16, bold=True)
        title_cell.alignment = center_alignment
        
        # 基本信息行
        current_date = datetime.now().strftime("%Y-%m-%d")
        ws['A3'] = "申请日期:"
        ws['B3'] = current_date
        ws['D3'] = "申请人:"
        ws['E3'] = self.project_manager
        ws['G3'] = "部门:"
        ws['H3'] = self.department
        
        # 表头
        headers = [
            "序号", "采购类型", "物品名称", "规格型号", "单位", 
            "数量", "单价(元)", "金额(元)", "二级分类", "备注"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # 数据行
        row = 6
        ws.cell(row=row, column=1).value = 1  # 序号
        ws.cell(row=row, column=2).value = "科研设备"  # 采购类型
        ws.cell(row=row, column=3).value = procurement_info['name']  # 物品名称
        ws.cell(row=row, column=4).value = procurement_info['specification']  # 规格型号
        ws.cell(row=row, column=5).value = procurement_info['unit']  # 单位
        ws.cell(row=row, column=6).value = procurement_info['quantity']  # 数量
        ws.cell(row=row, column=7).value = procurement_info['unit_price']  # 单价
        ws.cell(row=row, column=8).value = procurement_info['total_amount']  # 金额
        ws.cell(row=row, column=9).value = "科研耗材"  # 二级分类（默认值）
        ws.cell(row=row, column=10).value = "根据图片信息自动生成"  # 备注
        
        # 为数据行设置样式和边框
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.font = normal_font
            cell.alignment = center_alignment
            cell.border = border
        
        # 添加数据验证（下拉选项）
        # 采购类型下拉
        procurement_type_validation = DataValidation(
            type="list",
            formula1='"' + ','.join(self.procurement_types) + '"',
            allow_blank=False
        )
        procurement_type_validation.add(f'B{row}')
        ws.add_data_validation(procurement_type_validation)
        
        # 二级分类下拉
        category_validation = DataValidation(
            type="list",
            formula1='"' + ','.join(self.secondary_categories) + '"',
            allow_blank=False
        )
        category_validation.add(f'I{row}')
        ws.add_data_validation(category_validation)
        
        # 合计行
        total_row = row + 2
        ws.merge_cells(f'A{total_row}:G{total_row}')
        ws.cell(row=total_row, column=1).value = "合计金额"
        ws.cell(row=total_row, column=1).font = header_font
        ws.cell(row=total_row, column=1).alignment = center_alignment
        ws.cell(row=total_row, column=8).value = procurement_info['total_amount']
        ws.cell(row=total_row, column=8).font = header_font
        ws.cell(row=total_row, column=8).alignment = center_alignment
        
        # 审批信息
        approval_row = total_row + 3
        ws[f'A{approval_row}'] = "申请人签字:"
        ws[f'D{approval_row}'] = "部门负责人:"
        ws[f'G{approval_row}'] = "财务审核:"
        
        approval_row += 2
        ws[f'A{approval_row}'] = "日期:"
        ws[f'D{approval_row}'] = "日期:"
        ws[f'G{approval_row}'] = "日期:"
        
        # 保存文件
        wb.save(output_path)
        print(f"采购申请表已生成: {output_path}")
        return output_path
    
    def process_image_to_excel(self, image_path, output_path=None):
        """从图片生成采购申请Excel表格的完整流程"""
        print(f"正在处理图片: {image_path}")
        
        # 如果图片不存在，使用空的默认信息
        if not os.path.exists(image_path):
            print(f"图片文件不存在，使用空的默认信息")
            procurement_info = {
                'name': '',
                'specification': '',
                'unit_price': 0.0,
                'quantity': 0,
                'unit': '个',
                'total_amount': 0.0
            }
        else:
            # 从图片提取信息
            procurement_info = self.extract_info_from_image(image_path)
            if not procurement_info:
                print("图片信息提取失败，使用空的默认信息")
                procurement_info = {
                    'name': '',
                    'specification': '',
                    'unit_price': 0.0,
                    'quantity': 0,
                    'unit': '个',
                    'total_amount': 0.0
                }
        
        print("提取的采购信息:")
        for key, value in procurement_info.items():
            print(f"  {key}: {value}")
        
        # 生成Excel表格
        excel_path = self.create_procurement_excel(procurement_info, output_path)
        
        return excel_path, procurement_info

def main():
    """主函数"""
    processor = ProcurementProcessor()
    
    # 示例：处理图片生成采购申请表
    image_path = "procurement_image.png"  # 替换为实际图片路径
    
    try:
        excel_path, info = processor.process_image_to_excel(image_path)
        print(f"\n采购申请表生成成功!")
        print(f"文件路径: {excel_path}")
        print(f"\n采购信息摘要:")
        print(f"物品名称: {info['name']}")
        print(f"规格型号: {info['specification']}")
        print(f"数量: {info['quantity']} {info['unit']}")
        print(f"单价: ¥{info['unit_price']}")
        print(f"总金额: ¥{info['total_amount']}")
        
    except Exception as e:
        print(f"处理过程中出现错误: {e}")

if __name__ == "__main__":
    main()

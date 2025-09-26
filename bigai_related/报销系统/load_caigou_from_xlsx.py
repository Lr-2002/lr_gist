import pandas as pd
import os
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

def load_and_analyze_excel(file_path):
    """
    加载Excel文件并分析其结构
    """
    try:
        # 读取Excel文件
        df = pd.read_excel(file_path)
        
        print("Excel文件列名:")
        print(df.columns.tolist())
        print(f"\n总行数: {len(df)}")
        print(f"\n前5行数据:")
        print(df.head())
        
        # 检查是否有订单状态列
        status_columns = [col for col in df.columns if '状态' in col or 'status' in col.lower()]
        print(f"\n可能的状态列: {status_columns}")
        
        if status_columns:
            print(f"\n{status_columns[0]}列的唯一值:")
            print(df[status_columns[0]].unique())
        
        return df
        
    except Exception as e:
        print(f"读取Excel文件时出错: {e}")
        return None

def extract_successful_orders(df):
    """
    提取交易成功的订单数据
    """
    # 查找状态列
    status_column = None
    for col in df.columns:
        if '状态' in col or 'status' in col.lower():
            status_column = col
            break
    
    if not status_column:
        print("未找到订单状态列")
        return []
    
    # 筛选交易成功的订单
    successful_orders = df[df[status_column] == '交易成功']
    print(f"找到 {len(successful_orders)} 个交易成功的订单")
    
    # 根据订单号去重
    order_id_column = None
    for col in df.columns:
        if '订单号' in col or 'order' in col.lower():
            order_id_column = col
            break
    
    if order_id_column:
        before_dedup = len(successful_orders)
        successful_orders = successful_orders.drop_duplicates(subset=[order_id_column])
        after_dedup = len(successful_orders)
        print(f"根据订单号去重：{before_dedup} -> {after_dedup} 个订单")
    else:
        print("未找到订单号列，跳过去重")
    
    # 提取所需字段
    extracted_data = []
    
    # 定义字段映射（可能的列名）
    field_mapping = {
        '店铺名称': ['店铺名称', '商家', '店铺', '卖家'],
        '商品名称': ['商品名称', '商品', '产品名称', '标题'],
        '型号规格': ['型号规格', '规格', '型号', '规格型号'],
        '商品数量': ['商品数量', '数量', '购买数量'],
        '金额': [ '实付金额']
    }
    
    # 找到实际的列名
    actual_columns = {}
    for field, possible_names in field_mapping.items():
        for name in possible_names:
            if name in df.columns:
                actual_columns[field] = name
                break
    
    print(f"找到的列映射: {actual_columns}")
    
    # 提取数据
    for _, row in successful_orders.iterrows():
        order_data = {}
        for field, column_name in actual_columns.items():
            order_data[field] = row[column_name] if column_name in row else None
        extracted_data.append(order_data)
    
    return extracted_data

def export_to_excel(data, output_file):
    """
    将提取的数据导出到Excel文件
    """
    try:
        df_export = pd.DataFrame(data)
        df_export.to_excel(output_file, index=False, engine='openpyxl')
        print(f"\n数据已导出到: {output_file}")
    except Exception as e:
        print(f"导出Excel文件时出错: {e}")

def get_summary_statistics(data):
    """
    获取数据统计信息
    """
    if not data:
        return
    
    print(f"\n=== 数据统计 ===")
    print(f"交易成功订单总数: {len(data)}")
    
    # 统计店铺
    shops = [order.get('店铺名称', '') for order in data if order.get('店铺名称')]
    shop_counts = {}
    for shop in shops:
        shop_counts[shop] = shop_counts.get(shop, 0) + 1
    
    print(f"\n店铺统计:")
    for shop, count in sorted(shop_counts.items(), key=lambda x: x[1], reverse=True):
        print(f"  {shop}: {count}个订单")
    
    # 统计金额
    amounts = []
    for order in data:
        amount_str = order.get('金额', '0')
        if amount_str:
            amount_str = str(amount_str)
            # 使用正则表达式提取数字（包括小数点）
            numbers = re.findall(r'\d+\.?\d*', amount_str)
            if numbers:
                try:
                    amounts.append(float(numbers[0]))
                except ValueError:
                    continue
    
    if amounts:
        total_amount = sum(amounts)
        avg_amount = total_amount / len(amounts)
        print(f"\n金额统计:")
        print(f"  总金额: ￥{total_amount:.2f}")
        print(f"  平均金额: ￥{avg_amount:.2f}")
        print(f"  最高金额: ￥{max(amounts):.2f}")
        print(f"  最低金额: ￥{min(amounts):.2f}")

def generate_procurement_request(orders_data, output_file):
    """
    根据订单数据生成采购申请表
    """
    try:
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "采购申请表"
        
        # 写入标题"明细表1"
        ws.merge_cells('A1:G1')
        title_cell = ws.cell(row=1, column=1, value="明细表1")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 设置表头
        headers = ["物资一级分类", "物资名称", "规格型号", "单位", "数量", "估算单价/元", "估算总价/元"]
        
        # 写入表头（第2行）
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # 填入数据（从第3行开始）
        for row_idx, order in enumerate(orders_data, 3):
            # 物资一级分类 - 固定为"科研耗材"
            ws.cell(row=row_idx, column=1, value="科研耗材")
            
            # 物资名称 - 使用商品名称
            product_name = order.get('商品名称', '')
            ws.cell(row=row_idx, column=2, value=product_name)
            
            # 规格型号 - 也使用商品名称
            ws.cell(row=row_idx, column=3, value=product_name)
            
            # 单位 - 固定为"批"
            ws.cell(row=row_idx, column=4, value="批")
            
            # 数量 - 固定为1
            ws.cell(row=row_idx, column=5, value=1)
            
            # 估算单价和总价 - 使用订单金额
            amount_str = order.get('金额', '0')
            amount = 0
            if amount_str:
                # 转换为字符串处理
                amount_str = str(amount_str)
                # 使用正则表达式提取数字（包括小数点）
                numbers = re.findall(r'\d+\.?\d*', amount_str)
                if numbers:
                    try:
                        amount = float(numbers[0])
                    except ValueError:
                        amount = 0
            
            # 估算单价
            ws.cell(row=row_idx, column=6, value=amount)
            
            # 估算总价（数量为1，所以等于单价）
            ws.cell(row=row_idx, column=7, value=amount)
            
            # 为数据行添加边框
            for col in range(1, 8):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 调整列宽
        column_widths = [15, 30, 30, 8, 8, 15, 15]
        column_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
        for i, width in enumerate(column_widths):
            ws.column_dimensions[column_letters[i]].width = width
        
        # 保存文件
        wb.save(output_file)
        print(f"\n采购申请表已生成: {output_file}")
        
        # 显示生成的数据统计
        total_amount = 0
        for order in orders_data:
            amount_str = order.get('金额', '0')
            if amount_str:
                amount_str = str(amount_str)
                numbers = re.findall(r'\d+\.?\d*', amount_str)
                if numbers:
                    try:
                        total_amount += float(numbers[0])
                    except ValueError:
                        continue
        print(f"共生成 {len(orders_data)} 项采购申请")
        print(f"总金额: ￥{total_amount:.2f}")
        
    except Exception as e:
        print(f"生成采购申请表时出错: {e}")

def main():
    # Excel文件路径
    excel_file = "/Users/lr-2002/Documents/报销材料/9.26采购申请/订单数据.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"文件不存在: {excel_file}")
        return
    
    # 加载并分析Excel文件
    df = load_and_analyze_excel(excel_file)
    
    if df is not None:
        # 提取交易成功的订单
        successful_orders = extract_successful_orders(df)
        
        # 显示提取的数据
        print(f"\n提取到的交易成功订单数据:")
        for i, order in enumerate(successful_orders, 1):
            print(f"\n订单 {i}:")
            for field, value in order.items():
                print(f"  {field}: {value}")
        
        # 显示统计信息
        get_summary_statistics(successful_orders)
        
        # 导出数据到Excel
        if successful_orders:
            output_file = "/Users/lr-2002/project/lr_gist/bigai_related/报销系统/交易成功订单.xlsx"
            export_to_excel(successful_orders, output_file)
            
            # 生成采购申请表
            today = datetime.now().strftime("%Y%m%d")
            procurement_file = f"/Users/lr-2002/project/lr_gist/bigai_related/报销系统/{today}_采购申请表.xlsx"
            generate_procurement_request(successful_orders, procurement_file)
        
        return successful_orders

if __name__ == "__main__":
    main()